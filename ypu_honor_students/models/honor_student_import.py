import base64
import io
import re

from odoo import _, fields, models
from odoo.exceptions import UserError

_COLUMN_LABELS = [
    'Student Name *',
    'Student Number',
    'Faculty *',
    'Year *',
    'Study Year *',
    'Semester *',
    'GPA',
    'Honor Title',
    'Achievement Summary',
    'Publish on Website (Yes/No)',
    'Publicly Visible (Yes/No)',
    'Sequence',
]


class YpuHonorStudentImport(models.TransientModel):
    _name = 'ypu.honor.student.import'
    _description = 'Import Honor Students from Excel'

    excel_file = fields.Binary(string='Excel File (.xlsx)', required=True, attachment=False)
    filename = fields.Char(string='Filename')
    update_existing = fields.Boolean(
        string='Update existing students (match by Student Number, then Name)',
        default=False,
    )

    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft', readonly=True)
    result_created = fields.Integer(string='Created', readonly=True)
    result_updated = fields.Integer(string='Updated', readonly=True)
    result_skipped = fields.Integer(string='Skipped', readonly=True)
    result_log = fields.Text(string='Import Log', readonly=True)

    def action_download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise UserError(_('openpyxl is not available in this environment.'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Honor Students'

        req_fill = PatternFill(fill_type='solid', fgColor='C00000')
        hdr_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', wrap_text=True)

        for col_idx, label in enumerate(_COLUMN_LABELS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = hdr_font
            cell.fill = req_fill if label.endswith('*') else hdr_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = 28

        sample = [
            'Ahmad Al-Hassan',
            '202300145',
            'Faculty of Engineering',
            'Year 2',
            'Year 2',
            'Semester 1',
            '3.92',
            'Dean List',
            'Top ranked student in the faculty for this semester.',
            'Yes',
            'Yes',
            '10',
        ]
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = base64.b64encode(buf.read()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'honor_students_import_template.xlsx',
            'type': 'binary',
            'datas': data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    @staticmethod
    def _is_true(text):
        return str(text or '').strip().lower() in ('1', 'yes', 'true', 'y')

    @staticmethod
    def _parse_semester(text):
        raw = str(text or '').strip().lower()
        if raw in ('1', 'semester 1', 'sem 1', 's1'):
            return '1'
        if raw in ('2', 'semester 2', 'sem 2', 's2'):
            return '2'
        if raw in ('3', 'semester 3', 'sem 3', 's3'):
            return '3'
        return False

    @staticmethod
    def _parse_year_code(text):
        raw = str(text or '').strip().lower()
        if raw in ('1', 'year 1', 'y1'):
            return '1'
        if raw in ('2', 'year 2', 'y2'):
            return '2'
        if raw in ('3', 'year 3', 'y3'):
            return '3'
        if raw in ('4', 'year 4', 'y4'):
            return '4'
        if raw in ('5', 'year 5', 'y5'):
            return '5'
        match = re.search(r'([1-5])', raw)
        return match.group(1) if match else False

    def action_import(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_('openpyxl is not available in this environment.'))

        if not self.excel_file:
            raise UserError(_('Please upload an Excel file first.'))

        try:
            raw = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise UserError(_('Could not open the Excel file: %s') % exc)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        Student = self.env['ypu.honor.student']
        Faculty = self.env['ypu.honor.faculty']
        Year = self.env['ypu.honor.year']
        StudyYear = self.env['ypu.honor.study.year']

        created = updated = skipped = 0
        log_lines = []

        for row_num, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            def _val(idx, default=''):
                try:
                    v = row[idx]
                    return str(v).strip() if v is not None else default
                except (IndexError, TypeError):
                    return default

            name = _val(0)
            if not name:
                skipped += 1
                log_lines.append(f'Row {row_num}: skipped — Student Name is required.')
                continue

            faculty_name = _val(2)
            if not faculty_name:
                skipped += 1
                log_lines.append(f'Row {row_num}: skipped — Faculty is required for "{name}".')
                continue

            faculty = Faculty.search([('name', '=ilike', faculty_name)], limit=1)
            if not faculty:
                faculty = Faculty.create({'name': faculty_name})
                log_lines.append(f'Row {row_num}: created new faculty "{faculty_name}".')

            year_code = self._parse_year_code(_val(3))
            if not year_code:
                skipped += 1
                log_lines.append(f'Row {row_num}: skipped — invalid Year for "{name}".')
                continue

            year = Year.search([('code', '=', year_code)], limit=1)
            if not year:
                year = Year.create({'code': year_code, 'name': f'Year {year_code}'})
                log_lines.append(f'Row {row_num}: created missing year Year {year_code}.')

            study_year_code = self._parse_year_code(_val(4)) or year_code
            study_year = StudyYear.search([('code', '=', study_year_code)], limit=1)
            if not study_year:
                study_year = StudyYear.create({'code': study_year_code, 'name': f'Year {study_year_code}'})
                log_lines.append(f'Row {row_num}: created missing study year Year {study_year_code}.')

            semester = self._parse_semester(_val(5))
            if not semester:
                skipped += 1
                log_lines.append(f'Row {row_num}: skipped — invalid Semester for "{name}".')
                continue

            gpa_raw = _val(6)
            gpa = 0.0
            if gpa_raw:
                try:
                    gpa = float(gpa_raw)
                except ValueError:
                    skipped += 1
                    log_lines.append(f'Row {row_num}: skipped — invalid GPA for "{name}".')
                    continue

            vals = {
                'name': name,
                'student_number': _val(1) or False,
                'faculty_id': faculty.id,
                'year_id': year.id,
                'study_year_id': study_year.id,
                'semester': semester,
                'gpa': gpa,
                'honor_title': _val(7) or False,
                'achievement': _val(8) or False,
                'website_published': self._is_true(_val(9)) if _val(9) else True,
                'is_public': self._is_true(_val(10)) if _val(10) else True,
            }

            seq_raw = _val(11)
            if seq_raw and seq_raw.replace('.', '', 1).isdigit():
                vals['sequence'] = int(float(seq_raw))

            existing = False
            if self.update_existing:
                student_number = vals.get('student_number')
                if student_number:
                    existing = Student.search([('student_number', '=', student_number)], limit=1)
                if not existing:
                    existing = Student.search([('name', '=ilike', name)], limit=1)

            if existing:
                existing.write(vals)
                updated += 1
                log_lines.append(f'Row {row_num}: updated "{name}".')
            else:
                Student.create(vals)
                created += 1
                log_lines.append(f'Row {row_num}: created "{name}".')

        if not log_lines:
            log_lines.append('Import completed. No data rows found.')

        self.write({
            'result_created': created,
            'result_updated': updated,
            'result_skipped': skipped,
            'result_log': '\n'.join(log_lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ypu.honor.student.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
