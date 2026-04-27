import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Columns in the expected order (index corresponds to Excel column A=0, B=1, …)
_COLUMN_LABELS = [
    'Name *',
    'Department',
    'Subject',
    'Rank',
    'Category',
    'Position',
    'Email',
    'Phone',
    'LinkedIn URL',
    'Research Gate URL',
    'Bio / Summary (short)',
    'Available (Yes/No)',
    'Featured/Dean (Yes/No)',
    'Sequence',
    'Personal Information',
    'Education',
    'Career',
    'Administration',
    'Supervising',
    'Publications',
    'Courses',
]


class YpuTeacherImport(models.TransientModel):
    _name = 'ypu.teacher.import'
    _description = 'Import Teachers from Excel'

    excel_file = fields.Binary(string='Excel File (.xlsx)', required=True, attachment=False)
    filename = fields.Char(string='Filename')
    update_existing = fields.Boolean(
        string='Update existing teachers (match by name)',
        default=False,
        help='When enabled, rows whose Name matches an existing teacher will update '
             'that record instead of creating a duplicate.',
    )

    # ── Results (populated after import) ─────────────────────
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft', readonly=True)
    result_created = fields.Integer(string='Created', readonly=True)
    result_updated = fields.Integer(string='Updated', readonly=True)
    result_skipped = fields.Integer(string='Skipped', readonly=True)
    result_log = fields.Text(string='Import Log', readonly=True)

    # ── Template download ─────────────────────────────────────

    def action_download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise UserError(_("openpyxl is not available in this environment."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Teachers'

        req_fill = PatternFill(fill_type='solid', fgColor='C00000')
        hdr_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', wrap_text=True)

        for col_idx, label in enumerate(_COLUMN_LABELS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = hdr_font
            cell.fill = req_fill if label.endswith('*') else hdr_fill
            cell.alignment = center
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = 26

        # Sample row
        sample = [
            'Dr. John Smith', 'Computer Science', 'Data Structures',
            'Associate Professor', 'Engineering', 'Head of Department',
            'john@example.com', '+962-7-1234567',
            'https://linkedin.com/in/johnsmith',
            'https://researchgate.net/profile/johnsmith',
            'Expert in data structures, algorithms, and software engineering.',
            'Yes', 'No', '10',
            'Born 1970. Married.',
            'PhD - MIT 2000; MSc - Oxford 1996; BSc - Damascus 1992',
            '2010-present: Professor at YPU\n2005-2010: Lecturer at IUST',
            'Head of CS Department 2015-2020',
            'Supervised 5 PhD dissertations',
            '1. Smith J. "Title", Journal, 2020',
            'Data Structures | Algorithms | Database Systems',
        ]
        for col_idx, val in enumerate(sample, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = base64.b64encode(buf.read()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'teachers_import_template.xlsx',
            'type': 'binary',
            'datas': data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ── Import ────────────────────────────────────────────────

    def action_import(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is not available in this environment."))

        if not self.excel_file:
            raise UserError(_("Please upload an Excel file first."))

        try:
            raw = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise UserError(_("Could not open the Excel file: %s") % exc)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        Teacher = self.env['ypu.teacher']
        Category = self.env['ypu.teacher.category']
        Position = self.env['ypu.teacher.position']

        created = updated = skipped = 0
        log_lines = []

        for row_num, row in enumerate(rows, start=2):
            # Skip completely blank rows
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            def _val(idx, default=''):
                try:
                    v = row[idx]
                    return str(v).strip() if v is not None else default
                except (IndexError, TypeError):
                    return default

            def _html(text):
                """Convert plain-text cell to minimal HTML paragraphs if not already HTML."""
                if not text:
                    return ''
                text = text.strip()
                if text.startswith('<'):
                    return text
                # Convert newlines to <p> tags
                paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
                return ''.join(f'<p class="lead">{p}</p>' for p in paragraphs)

            name = _val(0)
            if not name:
                log_lines.append(f"Row {row_num}: skipped — Name is required.")
                skipped += 1
                continue

            vals = {
                'name': name,
                'department': _val(1) or False,
                'subject': _val(2) or False,
                'rank': _val(3) or False,
                'email': _val(6) or False,
                'phone': _val(7) or False,
                'linkedin_url': _val(8) or False,
                'research_gate': _val(9) or False,
                'bio': _html(_val(10)) or False,
                'available': _val(11).lower() not in ('no', 'false', '0'),
                'is_dean': _val(12).lower() in ('yes', 'true', '1'),
                'website_published': True,
                'is_public': True,
                'personal_info':  _html(_val(14)) or False,
                'education':      _html(_val(15)) or False,
                'career':         _html(_val(16)) or False,
                'administration': _html(_val(17)) or False,
                'supervising':    _html(_val(18)) or False,
                'publications':   _html(_val(19)) or False,
                'courses':        _html(_val(20)) or False,
            }

            seq_raw = _val(13)
            if seq_raw and seq_raw.replace('.', '', 1).isdigit():
                vals['sequence'] = int(float(seq_raw))

            # Category — create if new
            cat_name = _val(4)
            if cat_name:
                cat = Category.search([('name', '=ilike', cat_name)], limit=1)
                if not cat:
                    cat = Category.create({'name': cat_name})
                    log_lines.append(f"Row {row_num}: created new category '{cat_name}'.")
                vals['category_id'] = cat.id

            # Position — create if new
            pos_name = _val(5)
            if pos_name:
                pos = Position.search([('name', '=ilike', pos_name)], limit=1)
                if not pos:
                    pos = Position.create({'name': pos_name})
                    log_lines.append(f"Row {row_num}: created new position '{pos_name}'.")
                vals['position_id'] = pos.id

            # Update or create
            existing = False
            if self.update_existing:
                existing = Teacher.search([('name', '=ilike', name)], limit=1)

            if existing:
                existing.write(vals)
                existing._create_website_page()
                updated += 1
                log_lines.append(
                    f"Row {row_num}: updated '{name}' → {existing.link_url}"
                )
            else:
                teacher = Teacher.create(vals)
                teacher._create_website_page()
                created += 1
                log_lines.append(
                    f"Row {row_num}: created '{name}' → {teacher.link_url}"
                )

        if not log_lines:
            log_lines.append("Import completed. No data rows found.")

        self.write({
            'result_created': created,
            'result_updated': updated,
            'result_skipped': skipped,
            'result_log': '\n'.join(log_lines),
            'state': 'done',
        })

        # Re-open the same wizard to show results
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ypu.teacher.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
