import base64
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Columns in the expected order (index corresponds to Excel column A=0, B=1, …)
_COLUMN_LABELS = [
    'Plan Name *',
    'Plan Code',
    'Plan Subtitle',
    'Section Name *',
    'Section Subtitle',
    'Section Type',
    'Show Prerequisite Column (Yes/No)',
    'Show Description Column (Yes/No)',
    'Course Code',
    'Course Name (Arabic)',
    'Course Name (English)',
    'Theory Hours',
    'Practical Hours',
    'Credit Hours',
    'Prerequisite',
    'Course Description',
]

_SECTION_TYPE_MAP = {
    'category': 'category',
    'semester': 'semester',
    'elective': 'elective',
    'custom': 'custom',
}


class YpuStudyPlanImport(models.TransientModel):
    _name = 'ypu.study.plan.import'
    _description = 'Import Study Plans from Excel'

    excel_file = fields.Binary(string='Excel File (.xlsx)', required=True, attachment=False)
    filename = fields.Char(string='Filename')
    update_existing = fields.Boolean(
        string='Update existing plans (match by name)',
        default=False,
        help='When enabled, rows whose Plan Name matches an existing plan will update '
             'that plan\'s code and subtitle instead of leaving them unchanged.',
    )

    # ── Results (populated after import) ─────────────────────
    state = fields.Selection([('draft', 'Draft'), ('done', 'Done')], default='draft', readonly=True)
    result_created = fields.Integer(string='Plans Created', readonly=True)
    result_updated = fields.Integer(string='Plans Updated', readonly=True)
    result_skipped = fields.Integer(string='Rows Skipped', readonly=True)
    result_log = fields.Text(string='Import Log', readonly=True)

    # ── Template download ─────────────────────────────────────

    def action_download_template(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            raise UserError(_("openpyxl is not available in this environment."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Study Plans'

        req_fill = PatternFill(fill_type='solid', fgColor='C00000')
        hdr_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
        hdr_font = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center', wrap_text=True)

        for col_idx, label in enumerate(_COLUMN_LABELS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = hdr_font
            cell.fill = req_fill if label.endswith('*') else hdr_fill
            cell.alignment = center
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = 26

        sample_rows = [
            [
                'Faculty of Dentistry Plan', 'dentistry_2024', '2024-2025',
                'University Mandatory Requirements', '12 credit hours', 'category', 'Yes', 'No',
                'URQ 121', 'مبادئ الكيمياء', 'Principles of Chemistry', '2', '1', '3', '', '',
            ],
            [
                'Faculty of Dentistry Plan', '', '',
                'University Mandatory Requirements', '', '', '', '',
                'URQ 122', 'فيزياء عامة', 'General Physics', '2', '1', '3', 'URQ 121', '',
            ],
            [
                'Faculty of Dentistry Plan', '', '',
                'Year 1 - Term 1', '18 credit hours', 'semester', 'Yes', 'No',
                'DEN 101', 'تشريح الأسنان', 'Dental Anatomy', '3', '2', '4', '', '',
            ],
        ]
        for row_num, row_data in enumerate(sample_rows, start=2):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=row_num, column=col_idx, value=val)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        data = base64.b64encode(buf.read()).decode()

        attachment = self.env['ir.attachment'].create({
            'name': 'study_plans_import_template.xlsx',
            'type': 'binary',
            'datas': data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ── Import ────────────────────────────────────────────────

    def action_import(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("openpyxl is not available in this environment."))

        if not self.excel_file:
            raise UserError(_("Please upload an Excel file first."))

        try:
            raw = base64.b64decode(self.excel_file)
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        except Exception as exc:
            raise UserError(_("Could not open the Excel file: %s") % exc)

        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        Plan = self.env['ypu.study.plan']
        Section = self.env['ypu.study.plan.section']
        Course = self.env['ypu.study.plan.course']

        skipped = 0
        log_lines = []
        plans_created = set()
        plans_updated = set()

        # Caches to avoid redundant searches within the same import
        plan_cache = {}    # plan_name.lower() → plan record
        section_cache = {} # (plan.id, section_name.lower()) → section record

        for row_num, row in enumerate(rows, start=2):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            def _val(idx, default=''):
                try:
                    v = row[idx]
                    return str(v).strip() if v is not None else default
                except (IndexError, TypeError):
                    return default

            def _float(idx):
                try:
                    v = row[idx]
                    if v is None or str(v).strip() == '':
                        return 0.0
                    return float(v)
                except (IndexError, TypeError, ValueError):
                    return 0.0

            plan_name = _val(0)
            section_name = _val(3)

            if not plan_name:
                log_lines.append(f"Row {row_num}: skipped — Plan Name is required.")
                skipped += 1
                continue
            if not section_name:
                log_lines.append(f"Row {row_num}: skipped — Section Name is required.")
                skipped += 1
                continue

            # ── Get or create plan ────────────────────────────────
            plan_key = plan_name.lower()
            if plan_key not in plan_cache:
                existing_plan = Plan.search([('name', '=ilike', plan_name)], limit=1)
                if existing_plan and self.update_existing:
                    plan_vals = {}
                    if _val(1):
                        plan_vals['code'] = _val(1)
                    if _val(2):
                        plan_vals['subtitle'] = _val(2)
                    if plan_vals:
                        try:
                            existing_plan.write(plan_vals)
                        except Exception as e:
                            log_lines.append(f"Row {row_num}: warning updating plan '{plan_name}': {e}")
                    plan_cache[plan_key] = existing_plan
                    plans_updated.add(existing_plan.id)
                    log_lines.append(f"Row {row_num}: updated plan '{plan_name}'.")
                elif existing_plan:
                    plan_cache[plan_key] = existing_plan
                else:
                    plan_vals = {'name': plan_name}
                    if _val(1):
                        plan_vals['code'] = _val(1)
                    if _val(2):
                        plan_vals['subtitle'] = _val(2)
                    new_plan = Plan.create(plan_vals)
                    plan_cache[plan_key] = new_plan
                    plans_created.add(new_plan.id)
                    log_lines.append(f"Row {row_num}: created plan '{plan_name}'.")

            plan = plan_cache[plan_key]

            # ── Get or create section ─────────────────────────────
            section_key = (plan.id, section_name.lower())
            if section_key not in section_cache:
                existing_section = Section.search([
                    ('plan_id', '=', plan.id),
                    ('name', '=ilike', section_name),
                ], limit=1)
                if existing_section:
                    section_cache[section_key] = existing_section
                else:
                    sec_type_raw = _val(5).lower()
                    sec_type = _SECTION_TYPE_MAP.get(sec_type_raw, 'category')
                    section_vals = {
                        'plan_id': plan.id,
                        'name': section_name,
                        'subtitle': _val(4) or False,
                        'section_type': sec_type,
                        'show_prerequisite_column': _val(6).lower() not in ('no', 'false', '0'),
                        'show_description_column': _val(7).lower() in ('yes', 'true', '1'),
                    }
                    new_section = Section.create(section_vals)
                    section_cache[section_key] = new_section
                    log_lines.append(
                        f"Row {row_num}: created section '{section_name}' in plan '{plan_name}'."
                    )

            section = section_cache[section_key]

            # ── Create course if course data is present ────────────
            course_code = _val(8)
            course_name_ar = _val(9)
            course_name_en = _val(10)
            theory = _float(11)
            practical = _float(12)
            credit = _float(13)
            prerequisite = _val(14)
            description = _val(15)

            if not course_code and not course_name_ar and not course_name_en:
                continue  # Row defines plan/section only — no course data

            Course.create({
                'section_id': section.id,
                'code': course_code or False,
                'name_ar': course_name_ar or False,
                'name_en': course_name_en or False,
                'theory_hours': theory,
                'practical_hours': practical,
                'credit_hours': credit,
                'prerequisite': prerequisite or False,
                'description': description or False,
            })
            course_label = course_code or course_name_en or course_name_ar
            log_lines.append(
                f"Row {row_num}: added course '{course_label}' → '{section_name}'."
            )

        if not log_lines:
            log_lines.append("Import completed. No data rows found.")

        self.write({
            'result_created': len(plans_created),
            'result_updated': len(plans_updated),
            'result_skipped': skipped,
            'result_log': '\n'.join(log_lines),
            'state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ypu.study.plan.import',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
